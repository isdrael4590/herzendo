import io
import json
from datetime import datetime
from functools import wraps
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.core.exceptions import PermissionDenied
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.db.models import Q
from .models import Paciente
from .forms import PacienteForm
from .forms_usuario import UsuarioCrearForm, UsuarioEditarForm


def role_required(*grupos):
    def decorator(view_func):
        @wraps(view_func)
        def wrapped(request, *args, **kwargs):
            if not request.user.is_authenticated:
                from django.contrib.auth.views import redirect_to_login
                return redirect_to_login(request.get_full_path())
            if request.user.is_superuser or request.user.groups.filter(name__in=grupos).exists():
                return view_func(request, *args, **kwargs)
            raise PermissionDenied
        return wrapped
    return decorator


@login_required
def lista(request):
    q = request.GET.get('q', '')
    pacientes = Paciente.objects.all()
    if q:
        pacientes = pacientes.filter(
            Q(nombre__icontains=q) |
            Q(hcu__icontains=q) |
            Q(cirujano__icontains=q) |
            Q(procedimiento__icontains=q)
        )
    return render(request, 'pacientes/lista.html', {'pacientes': pacientes, 'q': q})


@role_required('administrador', 'medico')
def crear(request):
    if request.method == 'POST':
        form = PacienteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro guardado correctamente.')
            return redirect('lista')
    else:
        initial = {k: v for k, v in request.GET.items() if v}
        form = PacienteForm(initial=initial)
    return render(request, 'pacientes/form.html', {'form': form, 'titulo': 'Nuevo Registro'})


@role_required('administrador', 'medico')
def editar(request, pk):
    paciente = get_object_or_404(Paciente, pk=pk)
    form = PacienteForm(request.POST or None, instance=paciente)
    if form.is_valid():
        form.save()
        messages.success(request, 'Registro actualizado correctamente.')
        return redirect('lista')
    return render(request, 'pacientes/form.html', {'form': form, 'titulo': f'Editar — {paciente.nombre}', 'object': paciente})


@role_required('administrador')
def eliminar(request, pk):
    paciente = get_object_or_404(Paciente, pk=pk)
    if request.method == 'POST':
        paciente.delete()
        messages.success(request, 'Registro eliminado.')
        return redirect('lista')
    return render(request, 'pacientes/confirmar_eliminar.html', {'paciente': paciente})


@login_required
def exportar_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    q = request.GET.get('q', '')
    pacientes = Paciente.objects.all()
    if q:
        pacientes = pacientes.filter(
            Q(nombre__icontains=q) | Q(hcu__icontains=q) |
            Q(cirujano__icontains=q) | Q(procedimiento__icontains=q)
        )

    wb = openpyxl.Workbook()

    # ── Hoja 1: Datos Generales ──────────────────────────────────────────────
    _exportar_hoja(wb.active, 'Datos Generales', pacientes, [
        ('N°',                  lambda p, i: i),
        ('Nombre',              lambda p, i: p.nombre),
        ('HCU',                 lambda p, i: p.hcu),
        ('Fecha Nacimiento',    lambda p, i: p.fecha_nacimiento),
        ('Edad Dx',             lambda p, i: float(p.edad_diagnostico) if p.edad_diagnostico is not None else ''),
        ('Sexo',                lambda p, i: p.get_sexo_display()),
        ('Fecha Procedimiento', lambda p, i: p.fecha_procedimiento),
        ('Procedimiento',       lambda p, i: p.procedimiento),
        ('Cirujano',            lambda p, i: p.cirujano),
        ('Primera Cirugía',     lambda p, i: p.get_primera_cirugia_display()),
        ('Solo Biopsia',        lambda p, i: p.get_solo_biopsia_display()),
        ('Provincia',           lambda p, i: p.provincia_nacimiento),
        ('País Nacimiento',     lambda p, i: p.pais_nacimiento),
        ('Extranjero',          lambda p, i: p.get_extranjero_display()),
        ('Nivel Instrucción',   lambda p, i: p.get_nivel_instruccion_display()),
        ('Etnia',               lambda p, i: p.get_etnia_display()),
        ('Antec. Familiares',   lambda p, i: p.get_antecedentes_fam_display()),
        ('Estado',              lambda p, i: p.get_estado_display()),
        ('Registrado',          lambda p, i: p.creado.strftime('%d/%m/%Y') if p.creado else ''),
    ])

    # ── Hoja 2: Hormonas ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Hormonas')
    _exportar_hoja(ws2, 'Hormonas', pacientes, [
        ('N°',                  lambda p, i: i),
        ('Nombre',              lambda p, i: p.nombre),
        ('HCU',                 lambda p, i: p.hcu),
        ('GH',                  lambda p, i: float(p.gh) if p.gh is not None else ''),
        ('GH Ref.Mín',          lambda p, i: float(p.gh_ref_min) if p.gh_ref_min is not None else ''),
        ('GH Ref.Máx',          lambda p, i: float(p.gh_ref_max) if p.gh_ref_max is not None else ''),
        ('GH Interp.',          lambda p, i: p.get_gh_interp_display()),
        ('IGF-1',               lambda p, i: float(p.igf1) if p.igf1 is not None else ''),
        ('IGF-1 Ref.Mín',       lambda p, i: float(p.igf1_ref_min) if p.igf1_ref_min is not None else ''),
        ('IGF-1 Ref.Máx',       lambda p, i: float(p.igf1_ref_max) if p.igf1_ref_max is not None else ''),
        ('IGF-1 Index',         lambda p, i: float(p.igf1_index_elevado) if p.igf1_index_elevado is not None else ''),
        ('Acromegalia',         lambda p, i: p.get_acromegalia_display()),
        ('Tto. Octreótide',     lambda p, i: p.get_tto_previo_octretide_display()),
        ('LH',                  lambda p, i: float(p.lh) if p.lh is not None else ''),
        ('FSH',                 lambda p, i: float(p.fsh) if p.fsh is not None else ''),
        ('Estradiol',           lambda p, i: float(p.estradiol) if p.estradiol is not None else ''),
        ('Testosterona',        lambda p, i: float(p.testosterona) if p.testosterona is not None else ''),
        ('Hipogonadismo Fem.',  lambda p, i: p.get_hipogonadismo_fem_display()),
        ('Hipogonadismo Masc.', lambda p, i: p.get_hipogonadismo_masc_display()),
        ('Amenorrea',           lambda p, i: p.get_amenorrea_display()),
        ('Gonadotropinoma',     lambda p, i: p.get_gonadotropinoma_display()),
        ('TSH',                 lambda p, i: float(p.tsh) if p.tsh is not None else ''),
        ('T4L',                 lambda p, i: float(p.t4l) if p.t4l is not None else ''),
        ('Hipotiroidismo Cent.', lambda p, i: p.get_hipotiroidismo_central_display()),
        ('Tirotropinoma',       lambda p, i: p.get_tirotropinoma_display()),
        ('Cortisol',            lambda p, i: float(p.cortisol) if p.cortisol is not None else ''),
        ('ACTH',                lambda p, i: float(p.acth) if p.acth is not None else ''),
        ('Cushing',             lambda p, i: p.get_cushing_display()),
        ('Prolactina',          lambda p, i: float(p.prolactina) if p.prolactina is not None else ''),
        ('Prolactina Interp.',  lambda p, i: p.get_prolactina_interp_display()),
        ('Tto. Cabergolina',    lambda p, i: p.get_tto_previo_cabergolina_display()),
    ])

    # ── Hoja 3: Tumor e Histología ───────────────────────────────────────────
    ws3 = wb.create_sheet('Tumor e Histología')
    _exportar_hoja(ws3, 'Tumor e Histología', pacientes, [
        ('N°',                  lambda p, i: i),
        ('Nombre',              lambda p, i: p.nombre),
        ('HCU',                 lambda p, i: p.hcu),
        ('Déf. Vasopresina',    lambda p, i: p.get_deficit_vasopresina_display()),
        ('Apoplejía',           lambda p, i: p.get_apoplejia_display()),
        ('Cefalea',             lambda p, i: p.get_cefalea_display()),
        ('Alt. Campo Visual',   lambda p, i: p.get_alt_campo_visual_display()),
        ('Convulsiones',        lambda p, i: p.get_convulsiones_display()),
        ('Fístula LCR',         lambda p, i: p.get_fistula_lcr_display()),
        ('Seno Cavernoso',      lambda p, i: p.get_sind_seno_cavernoso_display()),
        ('Tamaño Tumor (mm)',   lambda p, i: float(p.tamano_tumor) if p.tamano_tumor is not None else ''),
        ('Interp. Tumor',       lambda p, i: p.get_tamano_tumor_interp_display()),
        ('Invasión Seno Cav.',  lambda p, i: p.get_invasion_seno_cav_display()),
        ('KNOSP',               lambda p, i: p.get_knosp_display()),
        ('Tipo Histológico',    lambda p, i: p.get_tipo_histologico_display()),
        ('Reticulina Dist.',    lambda p, i: p.get_reticulina_distorsionada_display()),
        ('Hormonas IHQ',        lambda p, i: p.hormonas_ihq),
        ('Ki-67 (%)',           lambda p, i: float(p.ki67) if p.ki67 is not None else ''),
        ('Ki-67 Interp.',       lambda p, i: p.get_ki67_interp_display()),
        ('Observaciones',       lambda p, i: p.observaciones),
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = datetime.now().strftime('%Y%m%d_%H%M')
    nombre_archivo = f'pacientes_hipofisarios_{fecha}.xlsx'
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return resp


def _exportar_hoja(ws, titulo, pacientes, columnas):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
    title_cell = ws.cell(row=1, column=1, value=titulo)
    title_cell.font = Font(bold=True, size=13, color='FFFFFF')
    title_cell.fill = PatternFill('solid', fgColor='1a56db')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    # Cabecera
    hdr_fill = PatternFill('solid', fgColor='d1e0ff')
    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, (label, _) in enumerate(columnas, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = Font(bold=True, size=9)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 28

    # Datos
    alt_fill = PatternFill('solid', fgColor='F7F9FF')
    for row_idx, p in enumerate(pacientes, start=1):
        excel_row = row_idx + 2
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, (_, fn) in enumerate(columnas, start=1):
            try:
                val = fn(p, row_idx)
            except Exception:
                val = ''
            cell = ws.cell(row=excel_row, column=col_idx, value=val if val != '' else None)
            cell.font = Font(size=9)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if fill:
                cell.fill = fill

    # Ancho automático
    for col_idx in range(1, len(columnas) + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    max_len = max(max_len, len(str(cell.value or '')))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)


@login_required
def exportar_csv(request):
    import csv

    q = request.GET.get('q', '')
    pacientes = Paciente.objects.all()
    if q:
        pacientes = pacientes.filter(
            Q(nombre__icontains=q) | Q(hcu__icontains=q) |
            Q(cirujano__icontains=q) | Q(procedimiento__icontains=q)
        )

    COLUMNAS = [
        # ── Datos Generales ──────────────────────────────────────────────────
        ('nombre',               lambda p: p.nombre),
        ('hcu',                  lambda p: p.hcu),
        ('fecha_nacimiento',     lambda p: p.fecha_nacimiento.strftime('%d/%m/%Y') if p.fecha_nacimiento else ''),
        ('edad_diagnostico',     lambda p: float(p.edad_diagnostico) if p.edad_diagnostico is not None else ''),
        ('sexo',                 lambda p: p.get_sexo_display()),
        ('fecha_procedimiento',  lambda p: p.fecha_procedimiento.strftime('%d/%m/%Y') if p.fecha_procedimiento else ''),
        ('procedimiento',        lambda p: p.procedimiento),
        ('cirujano',             lambda p: p.cirujano),
        ('primera_cirugia',      lambda p: p.get_primera_cirugia_display()),
        ('solo_biopsia',         lambda p: p.get_solo_biopsia_display()),
        ('provincia_nacimiento', lambda p: p.provincia_nacimiento),
        ('pais_nacimiento',      lambda p: p.pais_nacimiento),
        ('extranjero',           lambda p: p.get_extranjero_display()),
        ('nivel_instruccion',    lambda p: p.get_nivel_instruccion_display()),
        ('etnia',                lambda p: p.get_etnia_display()),
        ('antecedentes_fam',     lambda p: p.get_antecedentes_fam_display()),
        ('estado',               lambda p: p.get_estado_display()),
        # ── Somatotropo ──────────────────────────────────────────────────────
        ('gh',                   lambda p: float(p.gh) if p.gh is not None else ''),
        ('gh_ref_min',           lambda p: float(p.gh_ref_min) if p.gh_ref_min is not None else ''),
        ('gh_ref_max',           lambda p: float(p.gh_ref_max) if p.gh_ref_max is not None else ''),
        ('gh_interp',            lambda p: p.get_gh_interp_display()),
        ('igf1',                 lambda p: float(p.igf1) if p.igf1 is not None else ''),
        ('igf1_ref_min',         lambda p: float(p.igf1_ref_min) if p.igf1_ref_min is not None else ''),
        ('igf1_ref_max',         lambda p: float(p.igf1_ref_max) if p.igf1_ref_max is not None else ''),
        ('igf1_interp',          lambda p: p.get_igf1_interp_display()),
        ('igf1_index_elevado',   lambda p: float(p.igf1_index_elevado) if p.igf1_index_elevado is not None else ''),
        ('acromegalia',          lambda p: p.get_acromegalia_display()),
        ('tto_previo_octretide', lambda p: p.get_tto_previo_octretide_display()),
        # ── Gonadotropo ──────────────────────────────────────────────────────
        ('lh',                   lambda p: float(p.lh) if p.lh is not None else ''),
        ('lh_ref_min',           lambda p: float(p.lh_ref_min) if p.lh_ref_min is not None else ''),
        ('lh_ref_max',           lambda p: float(p.lh_ref_max) if p.lh_ref_max is not None else ''),
        ('lh_interp',            lambda p: p.get_lh_interp_display()),
        ('fsh',                  lambda p: float(p.fsh) if p.fsh is not None else ''),
        ('fsh_ref_min',          lambda p: float(p.fsh_ref_min) if p.fsh_ref_min is not None else ''),
        ('fsh_ref_max',          lambda p: float(p.fsh_ref_max) if p.fsh_ref_max is not None else ''),
        ('fsh_interp',           lambda p: p.get_fsh_interp_display()),
        ('estradiol',            lambda p: float(p.estradiol) if p.estradiol is not None else ''),
        ('estradiol_ref_min',    lambda p: float(p.estradiol_ref_min) if p.estradiol_ref_min is not None else ''),
        ('estradiol_ref_max',    lambda p: float(p.estradiol_ref_max) if p.estradiol_ref_max is not None else ''),
        ('estradiol_interp',     lambda p: p.get_estradiol_interp_display()),
        ('testosterona',         lambda p: float(p.testosterona) if p.testosterona is not None else ''),
        ('testosterona_ref_min', lambda p: float(p.testosterona_ref_min) if p.testosterona_ref_min is not None else ''),
        ('testosterona_ref_max', lambda p: float(p.testosterona_ref_max) if p.testosterona_ref_max is not None else ''),
        ('testosterona_interp',  lambda p: p.get_testosterona_interp_display()),
        ('hipogonadismo_fem',    lambda p: p.get_hipogonadismo_fem_display()),
        ('hipogonadismo_masc',   lambda p: p.get_hipogonadismo_masc_display()),
        ('amenorrea',            lambda p: p.get_amenorrea_display()),
        ('gonadotropinoma',      lambda p: p.get_gonadotropinoma_display()),
        # ── Tiroideo ─────────────────────────────────────────────────────────
        ('tsh',                  lambda p: float(p.tsh) if p.tsh is not None else ''),
        ('tsh_ref_min',          lambda p: float(p.tsh_ref_min) if p.tsh_ref_min is not None else ''),
        ('tsh_ref_max',          lambda p: float(p.tsh_ref_max) if p.tsh_ref_max is not None else ''),
        ('t4l',                  lambda p: float(p.t4l) if p.t4l is not None else ''),
        ('t4l_ref_min',          lambda p: float(p.t4l_ref_min) if p.t4l_ref_min is not None else ''),
        ('t4l_ref_max',          lambda p: float(p.t4l_ref_max) if p.t4l_ref_max is not None else ''),
        ('hipotiroidismo_central',lambda p: p.get_hipotiroidismo_central_display()),
        ('tirotropinoma',        lambda p: p.get_tirotropinoma_display()),
        # ── Corticotropo ─────────────────────────────────────────────────────
        ('cortisol',             lambda p: float(p.cortisol) if p.cortisol is not None else ''),
        ('cortisol_ref_min',     lambda p: float(p.cortisol_ref_min) if p.cortisol_ref_min is not None else ''),
        ('cortisol_ref_max',     lambda p: float(p.cortisol_ref_max) if p.cortisol_ref_max is not None else ''),
        ('cortisol_interp',      lambda p: p.get_cortisol_interp_display()),
        ('eje_corticotropo_suf', lambda p: p.get_eje_corticotropo_suf_display()),
        ('eje_corticotropo_ins', lambda p: p.get_eje_corticotropo_ins_display()),
        ('acth',                 lambda p: float(p.acth) if p.acth is not None else ''),
        ('acth_ref_min',         lambda p: float(p.acth_ref_min) if p.acth_ref_min is not None else ''),
        ('acth_ref_max',         lambda p: float(p.acth_ref_max) if p.acth_ref_max is not None else ''),
        ('acth_interp',          lambda p: p.get_acth_interp_display()),
        ('cushing',              lambda p: p.get_cushing_display()),
        # ── Prolactina ───────────────────────────────────────────────────────
        ('prolactina',           lambda p: float(p.prolactina) if p.prolactina is not None else ''),
        ('prolactina_ref_min',   lambda p: float(p.prolactina_ref_min) if p.prolactina_ref_min is not None else ''),
        ('prolactina_ref_max',   lambda p: float(p.prolactina_ref_max) if p.prolactina_ref_max is not None else ''),
        ('prolactina_interp',    lambda p: p.get_prolactina_interp_display()),
        ('tto_previo_cabergolina',lambda p: p.get_tto_previo_cabergolina_display()),
        # ── Síntomas ─────────────────────────────────────────────────────────
        ('deficit_vasopresina',  lambda p: p.get_deficit_vasopresina_display()),
        ('apoplejia',            lambda p: p.get_apoplejia_display()),
        ('cefalea',              lambda p: p.get_cefalea_display()),
        ('alt_campo_visual',     lambda p: p.get_alt_campo_visual_display()),
        ('convulsiones',         lambda p: p.get_convulsiones_display()),
        ('fistula_lcr',          lambda p: p.get_fistula_lcr_display()),
        ('sind_seno_cavernoso',  lambda p: p.get_sind_seno_cavernoso_display()),
        # ── Tumor ────────────────────────────────────────────────────────────
        ('tamano_tumor',         lambda p: float(p.tamano_tumor) if p.tamano_tumor is not None else ''),
        ('tamano_tumor_interp',  lambda p: p.get_tamano_tumor_interp_display()),
        ('invasion_seno_cav',    lambda p: p.get_invasion_seno_cav_display()),
        ('knosp',                lambda p: p.get_knosp_display()),
        ('tipo_histologico',     lambda p: p.get_tipo_histologico_display()),
        # ── Histología ───────────────────────────────────────────────────────
        ('reticulina_distorsionada', lambda p: p.get_reticulina_distorsionada_display()),
        ('hormonas_ihq',         lambda p: p.hormonas_ihq),
        ('ki67',                 lambda p: float(p.ki67) if p.ki67 is not None else ''),
        ('ki67_interp',          lambda p: p.get_ki67_interp_display()),
        ('observaciones',        lambda p: p.observaciones),
    ]

    fecha = datetime.now().strftime('%Y%m%d_%H%M')
    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="pacientes_{fecha}.csv"'
    resp.write('﻿')  # BOM para compatibilidad Excel

    writer = csv.writer(resp)
    writer.writerow([col for col, _ in COLUMNAS])
    for p in pacientes:
        fila = []
        for _, fn in COLUMNAS:
            try:
                fila.append(fn(p))
            except Exception:
                fila.append('')
        writer.writerow(fila)
    return resp


@login_required
def plantilla_csv(request):
    import csv
    # Mismas columnas que el export para que sirva de referencia
    CABECERAS = [
        'nombre', 'hcu', 'fecha_nacimiento', 'edad_diagnostico', 'sexo',
        'fecha_procedimiento', 'procedimiento', 'cirujano',
        'primera_cirugia', 'solo_biopsia', 'provincia_nacimiento',
        'pais_nacimiento', 'extranjero', 'nivel_instruccion', 'etnia',
        'antecedentes_fam', 'estado',
        'gh', 'gh_ref_min', 'gh_ref_max', 'gh_interp',
        'igf1', 'igf1_ref_min', 'igf1_ref_max', 'igf1_interp', 'igf1_index_elevado',
        'acromegalia', 'tto_previo_octretide',
        'lh', 'lh_ref_min', 'lh_ref_max', 'lh_interp',
        'fsh', 'fsh_ref_min', 'fsh_ref_max', 'fsh_interp',
        'estradiol', 'estradiol_ref_min', 'estradiol_ref_max', 'estradiol_interp',
        'testosterona', 'testosterona_ref_min', 'testosterona_ref_max', 'testosterona_interp',
        'hipogonadismo_fem', 'hipogonadismo_masc', 'amenorrea', 'gonadotropinoma',
        'tsh', 'tsh_ref_min', 'tsh_ref_max',
        't4l', 't4l_ref_min', 't4l_ref_max',
        'hipotiroidismo_central', 'tirotropinoma',
        'cortisol', 'cortisol_ref_min', 'cortisol_ref_max', 'cortisol_interp',
        'eje_corticotropo_suf', 'eje_corticotropo_ins',
        'acth', 'acth_ref_min', 'acth_ref_max', 'acth_interp', 'cushing',
        'prolactina', 'prolactina_ref_min', 'prolactina_ref_max', 'prolactina_interp',
        'tto_previo_cabergolina',
        'deficit_vasopresina', 'apoplejia', 'cefalea', 'alt_campo_visual',
        'convulsiones', 'fistula_lcr', 'sind_seno_cavernoso',
        'tamano_tumor', 'tamano_tumor_interp', 'invasion_seno_cav', 'knosp',
        'tipo_histologico', 'reticulina_distorsionada',
        'hormonas_ihq', 'ki67', 'ki67_interp', 'observaciones',
    ]
    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = 'attachment; filename="plantilla_importar.csv"'
    resp.write('﻿')
    writer = csv.writer(resp)
    writer.writerow(CABECERAS)
    # Fila de ejemplo con solo los campos básicos
    ejemplo = {
        'nombre': 'Ejemplo Paciente', 'hcu': 'HCU-001',
        'fecha_nacimiento': '15/03/1980', 'edad_diagnostico': '45',
        'sexo': 'Masculino', 'fecha_procedimiento': '10/01/2025',
        'procedimiento': 'Adenomectomía transesfenoidal', 'cirujano': 'Dr. Ejemplo',
        'primera_cirugia': 'Sí', 'solo_biopsia': 'No',
        'provincia_nacimiento': 'Pichincha', 'pais_nacimiento': 'Ecuador',
        'extranjero': 'No', 'nivel_instruccion': 'Superior Completa',
        'etnia': 'Mestizo', 'antecedentes_fam': 'No', 'estado': 'Pendiente',
    }
    writer.writerow([ejemplo.get(c, '') for c in CABECERAS])
    return resp


@role_required('administrador', 'medico')
def importar_csv(request):
    import csv
    import io
    from decimal import Decimal, InvalidOperation
    from .models import (SI_NO, SI_NO_NA, SEXO, INSTRUCCION, ETNIA, INTERP,
                         PROLACT_INTERP, TUMOR_INTERP, KI67_INTERP, TIPO_HISTOLOGICO, KNOSP)

    if request.method == 'GET':
        return render(request, 'pacientes/importar_csv.html')

    archivo = request.FILES.get('csv_file')
    if not archivo:
        messages.error(request, 'No se seleccionó ningún archivo.')
        return render(request, 'pacientes/importar_csv.html')
    if not archivo.name.lower().endswith('.csv'):
        messages.error(request, 'El archivo debe tener extensión .csv')
        return render(request, 'pacientes/importar_csv.html')

    try:
        contenido = archivo.read().decode('utf-8-sig')
    except UnicodeDecodeError:
        archivo.seek(0)
        contenido = archivo.read().decode('latin-1')

    # ── Helpers ──────────────────────────────────────────────────────────────
    def _rev(choices):
        """Mapeo display→valor interno, acepta también el valor interno directamente."""
        m = {}
        for val, lbl in choices:
            if val:
                m[lbl.lower()] = val   # "Sí" → "si"
                m[val.lower()] = val   # "si" → "si"  (ya era interno)
        return m

    rev_si_no    = _rev(SI_NO)
    rev_si_no_na = _rev(SI_NO_NA)
    rev_sexo     = _rev(SEXO)
    rev_instruc  = _rev(INSTRUCCION)
    rev_etnia    = _rev(ETNIA)
    rev_interp   = _rev(INTERP)
    rev_prolact  = _rev(PROLACT_INTERP)
    rev_tumor    = _rev(TUMOR_INTERP)
    rev_ki67     = _rev(KI67_INTERP)
    rev_histol   = _rev(TIPO_HISTOLOGICO)
    rev_estado   = {'pendiente': 'pendiente', 'completado': 'completado',
                    'Pendiente'.lower(): 'pendiente', 'Completado'.lower(): 'completado'}

    # Provincias: valor interno == display, solo normalizar capitalización
    from .models import PROVINCIAS
    prov_map = {v.lower(): v for v, _ in PROVINCIAS if v}

    def _s(row, col):
        return (row.get(col) or '').strip()

    def _date(row, col):
        val = _s(row, col)
        if not val:
            return None
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                pass
        return None

    def _dec(row, col):
        val = _s(row, col).replace(',', '.')
        if not val:
            return None
        try:
            return Decimal(val)
        except InvalidOperation:
            return None

    def _choice(row, col, mapping, default=''):
        return mapping.get(_s(row, col).lower(), default)

    # ── Mapeo completo columna → función ─────────────────────────────────────
    def _build_defaults(row):
        return {
            # Datos generales
            'fecha_nacimiento':         _date(row, 'fecha_nacimiento'),
            'edad_diagnostico':         _dec(row, 'edad_diagnostico'),
            'sexo':                     _choice(row, 'sexo', rev_sexo),
            'fecha_procedimiento':      _date(row, 'fecha_procedimiento'),
            'procedimiento':            _s(row, 'procedimiento'),
            'cirujano':                 _s(row, 'cirujano'),
            'primera_cirugia':          _choice(row, 'primera_cirugia', rev_si_no),
            'solo_biopsia':             _choice(row, 'solo_biopsia', rev_si_no),
            'provincia_nacimiento':     prov_map.get(_s(row, 'provincia_nacimiento').lower(), ''),
            'pais_nacimiento':          _s(row, 'pais_nacimiento'),
            'extranjero':               _choice(row, 'extranjero', rev_si_no),
            'nivel_instruccion':        _choice(row, 'nivel_instruccion', rev_instruc),
            'etnia':                    _choice(row, 'etnia', rev_etnia),
            'antecedentes_fam':         _choice(row, 'antecedentes_fam', rev_si_no),
            'estado':                   _choice(row, 'estado', rev_estado, 'pendiente'),
            # Somatotropo
            'gh':                       _dec(row, 'gh'),
            'gh_ref_min':               _dec(row, 'gh_ref_min'),
            'gh_ref_max':               _dec(row, 'gh_ref_max'),
            'igf1':                     _dec(row, 'igf1'),
            'igf1_ref_min':             _dec(row, 'igf1_ref_min'),
            'igf1_ref_max':             _dec(row, 'igf1_ref_max'),
            'igf1_index_elevado':       _dec(row, 'igf1_index_elevado'),
            'acromegalia':              _choice(row, 'acromegalia', rev_si_no),
            'tto_previo_octretide':     _choice(row, 'tto_previo_octretide', rev_si_no),
            # Gonadotropo
            'lh':                       _dec(row, 'lh'),
            'lh_ref_min':               _dec(row, 'lh_ref_min'),
            'lh_ref_max':               _dec(row, 'lh_ref_max'),
            'fsh':                      _dec(row, 'fsh'),
            'fsh_ref_min':              _dec(row, 'fsh_ref_min'),
            'fsh_ref_max':              _dec(row, 'fsh_ref_max'),
            'estradiol':                _dec(row, 'estradiol'),
            'estradiol_ref_min':        _dec(row, 'estradiol_ref_min'),
            'estradiol_ref_max':        _dec(row, 'estradiol_ref_max'),
            'testosterona':             _dec(row, 'testosterona'),
            'testosterona_ref_min':     _dec(row, 'testosterona_ref_min'),
            'testosterona_ref_max':     _dec(row, 'testosterona_ref_max'),
            'hipogonadismo_fem':        _choice(row, 'hipogonadismo_fem', rev_si_no_na),
            'hipogonadismo_masc':       _choice(row, 'hipogonadismo_masc', rev_si_no_na),
            'amenorrea':                _choice(row, 'amenorrea', rev_si_no),
            'gonadotropinoma':          _choice(row, 'gonadotropinoma', rev_si_no),
            # Tiroideo
            'tsh':                      _dec(row, 'tsh'),
            'tsh_ref_min':              _dec(row, 'tsh_ref_min'),
            'tsh_ref_max':              _dec(row, 'tsh_ref_max'),
            't4l':                      _dec(row, 't4l'),
            't4l_ref_min':              _dec(row, 't4l_ref_min'),
            't4l_ref_max':              _dec(row, 't4l_ref_max'),
            'hipotiroidismo_central':   _choice(row, 'hipotiroidismo_central', rev_si_no),
            'tirotropinoma':            _choice(row, 'tirotropinoma', rev_si_no),
            # Corticotropo
            'cortisol':                 _dec(row, 'cortisol'),
            'cortisol_ref_min':         _dec(row, 'cortisol_ref_min'),
            'cortisol_ref_max':         _dec(row, 'cortisol_ref_max'),
            'eje_corticotropo_suf':     _choice(row, 'eje_corticotropo_suf', rev_si_no),
            'eje_corticotropo_ins':     _choice(row, 'eje_corticotropo_ins', rev_si_no),
            'acth':                     _dec(row, 'acth'),
            'acth_ref_min':             _dec(row, 'acth_ref_min'),
            'acth_ref_max':             _dec(row, 'acth_ref_max'),
            'cushing':                  _choice(row, 'cushing', rev_si_no),
            # Prolactina
            'prolactina':               _dec(row, 'prolactina'),
            'prolactina_ref_min':       _dec(row, 'prolactina_ref_min'),
            'prolactina_ref_max':       _dec(row, 'prolactina_ref_max'),
            'tto_previo_cabergolina':   _choice(row, 'tto_previo_cabergolina', rev_si_no),
            # Síntomas
            'deficit_vasopresina':      _choice(row, 'deficit_vasopresina', rev_si_no),
            'apoplejia':                _choice(row, 'apoplejia', rev_si_no),
            'cefalea':                  _choice(row, 'cefalea', rev_si_no),
            'alt_campo_visual':         _choice(row, 'alt_campo_visual', rev_si_no),
            'convulsiones':             _choice(row, 'convulsiones', rev_si_no),
            'fistula_lcr':              _choice(row, 'fistula_lcr', rev_si_no),
            'sind_seno_cavernoso':      _choice(row, 'sind_seno_cavernoso', rev_si_no),
            # Tumor
            'tamano_tumor':             _dec(row, 'tamano_tumor'),
            'invasion_seno_cav':        _choice(row, 'invasion_seno_cav', rev_si_no),
            'knosp':                    _s(row, 'knosp'),
            'tipo_histologico':         _choice(row, 'tipo_histologico', rev_histol),
            # Histología
            'reticulina_distorsionada': _choice(row, 'reticulina_distorsionada', rev_si_no),
            'hormonas_ihq':             _s(row, 'hormonas_ihq'),
            'ki67':                     _dec(row, 'ki67'),
            'observaciones':            _s(row, 'observaciones'),
        }

    reader = csv.DictReader(io.StringIO(contenido))
    creados = actualizados = 0
    errores = []

    for i, row in enumerate(reader, start=2):
        nombre = _s(row, 'nombre')
        if not nombre:
            errores.append(f'Fila {i}: "nombre" es obligatorio.')
            continue
        hcu = _s(row, 'hcu')
        try:
            defaults = _build_defaults(row)
            if hcu:
                _, created = Paciente.objects.update_or_create(
                    hcu=hcu, defaults={**defaults, 'nombre': nombre}
                )
            else:
                Paciente.objects.create(nombre=nombre, **defaults)
                created = True
            if created:
                creados += 1
            else:
                actualizados += 1
        except Exception as e:
            errores.append(f'Fila {i} ({nombre}): {e}')

    resumen = []
    if creados:
        resumen.append(f'{creados} registro(s) creado(s)')
    if actualizados:
        resumen.append(f'{actualizados} registro(s) actualizado(s)')
    if resumen:
        messages.success(request, ', '.join(resumen) + '.')
    for err in errores:
        messages.warning(request, err)

    return redirect('lista')


@role_required('administrador', 'medico')
def importar_excel(request):
    if request.method == 'GET':
        return render(request, 'pacientes/importar_excel.html')

    archivo = request.FILES.get('excel_file')
    if not archivo:
        messages.error(request, 'No se seleccionó ningún archivo.')
        return render(request, 'pacientes/importar_excel.html')
    if not archivo.name.lower().endswith(('.xlsx', '.xls')):
        messages.error(request, 'El archivo debe ser .xlsx o .xls')
        return render(request, 'pacientes/importar_excel.html')

    import openpyxl
    from decimal import Decimal, InvalidOperation
    from .models import SI_NO, SI_NO_NA, SEXO, INSTRUCCION, ETNIA, INTERP
    from .models import PROLACT_INTERP, TUMOR_INTERP, KI67_INTERP, TIPO_HISTOLOGICO

    def _rev(choices):
        return {label.strip().lower(): val for val, label in choices if val}

    rev_si_no    = _rev(SI_NO)
    rev_si_no_na = _rev(SI_NO_NA)
    rev_sexo     = _rev(SEXO)
    rev_instruc  = _rev(INSTRUCCION)
    rev_etnia    = _rev(ETNIA)
    rev_interp   = _rev(INTERP)
    rev_prolact  = _rev(PROLACT_INTERP)
    rev_tumor    = _rev(TUMOR_INTERP)
    rev_ki67     = _rev(KI67_INTERP)
    rev_histol   = _rev(TIPO_HISTOLOGICO)

    def _s(v):
        return '' if v is None else str(v).strip()

    def _date(v):
        if v is None:
            return None
        if hasattr(v, 'year'):
            return v.date() if hasattr(v, 'date') and callable(v.date) else v
        s = _s(v)
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
        return None

    def _dec(v):
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return Decimal(str(v))
        s = _s(v).replace(',', '.')
        if not s:
            return None
        try:
            return Decimal(s)
        except InvalidOperation:
            return None

    def _r(mapping, v):
        return mapping.get(_s(v).lower(), '')

    HOJA1 = {
        'Nombre':              ('nombre',                lambda v: _s(v)),
        'HCU':                 ('hcu',                   lambda v: _s(v)),
        'Fecha Nacimiento':    ('fecha_nacimiento',       _date),
        'Edad Dx':             ('edad_diagnostico',       _dec),
        'Sexo':                ('sexo',                  lambda v: _r(rev_sexo, v)),
        'Fecha Procedimiento': ('fecha_procedimiento',    _date),
        'Procedimiento':       ('procedimiento',          lambda v: _s(v)),
        'Cirujano':            ('cirujano',               lambda v: _s(v)),
        'Primera Cirugía':     ('primera_cirugia',       lambda v: _r(rev_si_no, v)),
        'Solo Biopsia':        ('solo_biopsia',           lambda v: _r(rev_si_no, v)),
        'Provincia':           ('provincia_nacimiento',   lambda v: _s(v)),
        'País Nacimiento':     ('pais_nacimiento',        lambda v: _s(v)),
        'Extranjero':          ('extranjero',             lambda v: _r(rev_si_no, v)),
        'Nivel Instrucción':   ('nivel_instruccion',     lambda v: _r(rev_instruc, v)),
        'Etnia':               ('etnia',                  lambda v: _r(rev_etnia, v)),
        'Antec. Familiares':   ('antecedentes_fam',      lambda v: _r(rev_si_no, v)),
        'Estado':              ('estado',                 lambda v: _s(v).lower()),
    }
    HOJA2 = {
        'GH':                   ('gh',                    _dec),
        'GH Ref.Mín':           ('gh_ref_min',            _dec),
        'GH Ref.Máx':           ('gh_ref_max',            _dec),
        'GH Interp.':           ('gh_interp',             lambda v: _r(rev_interp, v)),
        'IGF-1':                ('igf1',                  _dec),
        'IGF-1 Ref.Mín':        ('igf1_ref_min',          _dec),
        'IGF-1 Ref.Máx':        ('igf1_ref_max',          _dec),
        'IGF-1 Index':          ('igf1_index_elevado',    _dec),
        'Acromegalia':          ('acromegalia',            lambda v: _r(rev_si_no, v)),
        'Tto. Octreótide':      ('tto_previo_octretide',  lambda v: _r(rev_si_no, v)),
        'LH':                   ('lh',                    _dec),
        'FSH':                  ('fsh',                   _dec),
        'Estradiol':            ('estradiol',              _dec),
        'Testosterona':         ('testosterona',           _dec),
        'Hipogonadismo Fem.':   ('hipogonadismo_fem',     lambda v: _r(rev_si_no_na, v)),
        'Hipogonadismo Masc.':  ('hipogonadismo_masc',    lambda v: _r(rev_si_no_na, v)),
        'Amenorrea':            ('amenorrea',              lambda v: _r(rev_si_no, v)),
        'Gonadotropinoma':      ('gonadotropinoma',        lambda v: _r(rev_si_no, v)),
        'TSH':                  ('tsh',                   _dec),
        'T4L':                  ('t4l',                   _dec),
        'Hipotiroidismo Cent.': ('hipotiroidismo_central', lambda v: _r(rev_si_no, v)),
        'Tirotropinoma':        ('tirotropinoma',          lambda v: _r(rev_si_no, v)),
        'Cortisol':             ('cortisol',               _dec),
        'ACTH':                 ('acth',                   _dec),
        'Cushing':              ('cushing',                lambda v: _r(rev_si_no, v)),
        'Prolactina':           ('prolactina',             _dec),
        'Prolactina Interp.':   ('prolactina_interp',     lambda v: _r(rev_prolact, v)),
        'Tto. Cabergolina':     ('tto_previo_cabergolina', lambda v: _r(rev_si_no, v)),
    }
    HOJA3 = {
        'Déf. Vasopresina':    ('deficit_vasopresina',        lambda v: _r(rev_si_no, v)),
        'Apoplejía':           ('apoplejia',                   lambda v: _r(rev_si_no, v)),
        'Cefalea':             ('cefalea',                     lambda v: _r(rev_si_no, v)),
        'Alt. Campo Visual':   ('alt_campo_visual',            lambda v: _r(rev_si_no, v)),
        'Convulsiones':        ('convulsiones',                lambda v: _r(rev_si_no, v)),
        'Fístula LCR':         ('fistula_lcr',                 lambda v: _r(rev_si_no, v)),
        'Seno Cavernoso':      ('sind_seno_cavernoso',         lambda v: _r(rev_si_no, v)),
        'Tamaño Tumor (mm)':   ('tamano_tumor',                _dec),
        'Interp. Tumor':       ('tamano_tumor_interp',        lambda v: _r(rev_tumor, v)),
        'Invasión Seno Cav.':  ('invasion_seno_cav',          lambda v: _r(rev_si_no, v)),
        'KNOSP':               ('knosp',                       lambda v: _s(v)),
        'Tipo Histológico':    ('tipo_histologico',            lambda v: _r(rev_histol, v)),
        'Reticulina Dist.':    ('reticulina_distorsionada',    lambda v: _r(rev_si_no, v)),
        'Hormonas IHQ':        ('hormonas_ihq',                lambda v: _s(v)),
        'Ki-67 (%)':           ('ki67',                        _dec),
        'Ki-67 Interp.':       ('ki67_interp',                lambda v: _r(rev_ki67, v)),
        'Observaciones':       ('observaciones',               lambda v: _s(v)),
    }

    def _leer_hoja(ws, col_map, acumulado):
        filas = list(ws.iter_rows(min_row=2, values_only=True))
        if not filas:
            return
        cabecera = {i: _s(v) for i, v in enumerate(filas[0]) if v is not None}
        for fila in filas[1:]:
            if not fila or fila[0] is None:
                continue
            try:
                idx = int(fila[0])
            except (ValueError, TypeError):
                continue
            entry = acumulado.setdefault(idx, {})
            hcu_actual = entry.get('hcu', '')
            for i, val in enumerate(fila):
                col_label = cabecera.get(i, '')
                if col_label == 'HCU' and not hcu_actual:
                    entry['hcu'] = _s(val)
                if col_label in col_map:
                    field, fn = col_map[col_label]
                    try:
                        entry[field] = fn(val)
                    except Exception:
                        pass

    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
    except Exception as e:
        messages.error(request, f'No se pudo abrir el archivo: {e}')
        return render(request, 'pacientes/importar_excel.html')

    acumulado = {}
    hojas = wb.sheetnames
    if len(hojas) >= 1:
        _leer_hoja(wb[hojas[0]], HOJA1, acumulado)
    if len(hojas) >= 2:
        _leer_hoja(wb[hojas[1]], HOJA2, acumulado)
    if len(hojas) >= 3:
        _leer_hoja(wb[hojas[2]], HOJA3, acumulado)
    wb.close()

    creados = actualizados = 0
    errores = []

    for idx, datos in sorted(acumulado.items()):
        nombre = datos.pop('nombre', '') or ''
        hcu    = datos.pop('hcu', '') or ''
        if not nombre:
            errores.append(f'Fila {idx + 2}: "Nombre" vacío, se omite.')
            continue
        try:
            if hcu:
                _, created = Paciente.objects.update_or_create(
                    hcu=hcu, defaults={**datos, 'nombre': nombre}
                )
            else:
                Paciente.objects.create(nombre=nombre, **datos)
                created = True
            if created:
                creados += 1
            else:
                actualizados += 1
        except Exception as e:
            errores.append(f'Fila {idx + 2} ({nombre}): {e}')

    resumen = []
    if creados:
        resumen.append(f'{creados} registro(s) creado(s)')
    if actualizados:
        resumen.append(f'{actualizados} registro(s) actualizado(s)')
    if resumen:
        messages.success(request, ', '.join(resumen) + '.')
    for err in errores:
        messages.warning(request, err)

    return redirect('lista')


@role_required('administrador', 'medico')
@require_POST
def cambiar_estado(request, pk):
    paciente = get_object_or_404(Paciente, pk=pk)
    nuevo = request.POST.get('estado')
    if nuevo not in ('pendiente', 'completado'):
        return JsonResponse({'error': 'Estado inválido'}, status=400)
    paciente.estado = nuevo
    paciente.save(update_fields=['estado', 'actualizado'])
    return JsonResponse({'estado': paciente.estado, 'label': paciente.get_estado_display()})


@role_required('administrador', 'medico')
def subir_pdf(request):
    return render(request, 'pacientes/subir_pdf.html')


@role_required('administrador', 'medico')
@require_POST
def procesar_pdf(request):
    from .pdf_extractor import extraer_todo

    archivo = request.FILES.get('pdf')
    if not archivo:
        return JsonResponse({'error': 'No se recibió ningún archivo.'}, status=400)
    if not archivo.name.lower().endswith('.pdf'):
        return JsonResponse({'error': 'El archivo debe ser un PDF.'}, status=400)
    if archivo.size > 50 * 1024 * 1024:
        return JsonResponse({'error': 'El archivo no debe superar 50 MB.'}, status=400)

    try:
        texto, campos, paginas, gemini_status = extraer_todo(archivo)
        return JsonResponse({
            'campos': campos,
            'paginas': paginas,
            'gemini_status': gemini_status,
            'texto_crudo': texto[:4000],
            'total_campos': len(campos),
        })
    except Exception as e:
        return JsonResponse({'error': f'No se pudo procesar el PDF: {str(e)}'}, status=500)


@login_required
def imprimir(request, pk):
    paciente = get_object_or_404(Paciente, pk=pk)
    return render(request, 'pacientes/imprimir.html', {'paciente': paciente})


# ── Gestión de usuarios (solo administrador) ──────────────────────────────────

@role_required('administrador')
def usuarios_lista(request):
    usuarios = User.objects.exclude(is_superuser=True).order_by('username')
    return render(request, 'pacientes/usuarios_lista.html', {'usuarios': usuarios})


def _sync_ver_codificado(usuario, activo):
    grp, _ = Group.objects.get_or_create(name='ver_codificado')
    if activo:
        usuario.groups.add(grp)
    else:
        usuario.groups.remove(grp)


@role_required('administrador')
def usuario_crear(request):
    form = UsuarioCrearForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        cd = form.cleaned_data
        u = User.objects.create_user(
            username=cd['username'],
            password=cd['password1'],
            first_name=cd.get('first_name', ''),
            last_name=cd.get('last_name', ''),
            email=cd.get('email', ''),
        )
        grupos = []
        if cd['rol']:
            grp_rol, _ = Group.objects.get_or_create(name=cd['rol'])
            grupos.append(grp_rol)
        if cd.get('ver_codificado'):
            grp, _ = Group.objects.get_or_create(name='ver_codificado')
            grupos.append(grp)
        if cd.get('ver_estadisticas'):
            grp, _ = Group.objects.get_or_create(name='estadisticas')
            grupos.append(grp)
        u.groups.set(grupos)
        messages.success(request, f'Usuario "{u.username}" creado correctamente.')
        return redirect('usuarios_lista')
    return render(request, 'pacientes/usuario_form.html', {
        'form': form, 'titulo': 'Nuevo Usuario', 'es_edicion': False,
    })


@role_required('administrador')
def usuario_editar(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    grupos_actuales = list(usuario.groups.values_list('name', flat=True))
    rol_actual = next((g for g in grupos_actuales if g in ('administrador', 'medico', 'visualizador')), '')
    tiene_codificado   = 'ver_codificado' in grupos_actuales
    tiene_estadisticas = 'estadisticas'   in grupos_actuales
    initial = {
        'username':         usuario.username,
        'first_name':       usuario.first_name,
        'last_name':        usuario.last_name,
        'email':            usuario.email,
        'rol':              rol_actual,
        'ver_codificado':   tiene_codificado,
        'ver_estadisticas': tiene_estadisticas,
        'is_active':        usuario.is_active,
    }
    form = UsuarioEditarForm(request.POST or None, usuario=usuario, initial=initial)
    if request.method == 'POST' and form.is_valid():
        cd = form.cleaned_data
        usuario.username   = cd['username']
        usuario.first_name = cd.get('first_name', '')
        usuario.last_name  = cd.get('last_name', '')
        usuario.email      = cd.get('email', '')
        usuario.is_active  = cd.get('is_active', True)
        if cd.get('password1'):
            usuario.set_password(cd['password1'])
        usuario.save()
        grupos = []
        if cd['rol']:
            grp_rol, _ = Group.objects.get_or_create(name=cd['rol'])
            grupos.append(grp_rol)
        if cd.get('ver_codificado'):
            grp, _ = Group.objects.get_or_create(name='ver_codificado')
            grupos.append(grp)
        if cd.get('ver_estadisticas'):
            grp, _ = Group.objects.get_or_create(name='estadisticas')
            grupos.append(grp)
        usuario.groups.set(grupos)
        messages.success(request, f'Usuario "{usuario.username}" actualizado.')
        return redirect('usuarios_lista')
    return render(request, 'pacientes/usuario_form.html', {
        'form': form, 'titulo': f'Editar — {usuario.username}',
        'es_edicion': True, 'usuario': usuario,
    })


@role_required('administrador')
def usuario_eliminar(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    if usuario.is_superuser:
        messages.error(request, 'No se puede eliminar a un superusuario.')
        return redirect('usuarios_lista')
    if request.method == 'POST':
        nombre = usuario.username
        usuario.delete()
        messages.success(request, f'Usuario "{nombre}" eliminado.')
        return redirect('usuarios_lista')


# ── Datos codificados (solo administrador) ────────────────────────────────────

# Codificación numérica de todas las variables categóricas
_CODIGOS = {
    'sexo':                 {'': '', 'M': 1, 'F': 2},
    'primera_cirugia':      {'': '', 'si': 1, 'no': 2},
    'solo_biopsia':         {'': '', 'si': 1, 'no': 2},
    'extranjero':           {'': '', 'si': 1, 'no': 2},
    'provincia_nacimiento': {
        '': '',
        'Azuay': 1, 'Bolívar': 2, 'Cañar': 3, 'Carchi': 4,
        'Chimborazo': 5, 'Cotopaxi': 6, 'El Oro': 7, 'Esmeraldas': 8,
        'Galápagos': 9, 'Guayas': 10, 'Imbabura': 11, 'Loja': 12,
        'Los Ríos': 13, 'Manabí': 14, 'Morona Santiago': 15, 'Napo': 16,
        'Orellana': 17, 'Pastaza': 18, 'Pichincha': 19, 'Santa Elena': 20,
        'Santo Domingo de los Tsáchilas': 21, 'Sucumbíos': 22,
        'Tungurahua': 23, 'Zamora Chinchipe': 24,
    },
    'nivel_instruccion': {
        '': '', 'ninguna': 1, 'primaria_completa': 2, 'primaria_incompleta': 3,
        'secundaria_completa': 4, 'secundaria_incompleta': 5,
        'superior_completa': 6, 'superior_incompleta': 7,
    },
    'etnia': {
        '': '', 'mestizo': 1, 'indigena': 2, 'afroecuatoriano': 3,
        'blanco': 4, 'montubio': 5, 'otro': 6, 'no_dato': 7,
    },
    'antecedentes_fam':       {'': '', 'si': 1, 'no': 2},
    'estado':                 {'pendiente': 0, 'completado': 1},
    'gh_interp':              {'': '', 'normal': 1, 'elevado': 2, 'bajo': 3, 'no_det': 4},
    'igf1_interp':            {'': '', 'normal': 1, 'elevado': 2, 'bajo': 3, 'no_det': 4},
    'acromegalia':            {'': '', 'si': 1, 'no': 2},
    'tto_previo_octretide':   {'': '', 'si': 1, 'no': 2},
    'lh_interp':              {'': '', 'normal': 1, 'elevado': 2, 'bajo': 3, 'no_det': 4},
    'fsh_interp':             {'': '', 'normal': 1, 'elevado': 2, 'bajo': 3, 'no_det': 4},
    'estradiol_interp':       {'': '', 'normal': 1, 'elevado': 2, 'bajo': 3, 'no_det': 4},
    'testosterona_interp':    {'': '', 'normal': 1, 'elevado': 2, 'bajo': 3, 'no_det': 4},
    'hipogonadismo_fem':      {'': '', 'si': 1, 'no': 2, 'na': 3},
    'hipogonadismo_masc':     {'': '', 'si': 1, 'no': 2, 'na': 3},
    'amenorrea':              {'': '', 'si': 1, 'no': 2},
    'gonadotropinoma':        {'': '', 'si': 1, 'no': 2},
    'hipotiroidismo_central': {'': '', 'si': 1, 'no': 2},
    'tirotropinoma':          {'': '', 'si': 1, 'no': 2},
    'cortisol_interp':        {'': '', 'normal': 1, 'elevado': 2, 'bajo': 3, 'no_det': 4},
    'eje_corticotropo_suf':   {'': '', 'si': 1, 'no': 2},
    'eje_corticotropo_ins':   {'': '', 'si': 1, 'no': 2},
    'acth_interp':            {'': '', 'normal': 1, 'elevado': 2, 'bajo': 3, 'no_det': 4},
    'cushing':                {'': '', 'si': 1, 'no': 2},
    'prolactina_interp':      {'': '', 'normal': 1, 'baja': 2, 'elevada': 3, 'no_dato': 4},
    'tto_previo_cabergolina': {'': '', 'si': 1, 'no': 2},
    'deficit_vasopresina':    {'': '', 'si': 1, 'no': 2},
    'apoplejia':              {'': '', 'si': 1, 'no': 2},
    'cefalea':                {'': '', 'si': 1, 'no': 2},
    'alt_campo_visual':       {'': '', 'si': 1, 'no': 2},
    'convulsiones':           {'': '', 'si': 1, 'no': 2},
    'fistula_lcr':            {'': '', 'si': 1, 'no': 2},
    'sind_seno_cavernoso':    {'': '', 'si': 1, 'no': 2},
    'tamano_tumor_interp':    {'': '', 'microadenoma': 1, 'macroadenoma': 2, 'gigante': 3},
    'invasion_seno_cav':      {'': '', 'si': 1, 'no': 2},
    'knosp':                  {'': '', '0': 0, '1': 1, '2': 2, '3': 3, '4': 4},
    'tipo_histologico':       {
        '': '', 'prolactinoma': 1, 'somatotropinoma': 2, 'corticotropinoma': 3,
        'gonadotropinoma': 4, 'tirotropinoma': 5, 'carcinoma': 6, 'No_funcionante': 7,
    },
    'reticulina_distorsionada': {'': '', 'si': 1, 'no': 2},
    'ki67_interp': {'': '', 'muy_bajo': 1, 'bajo': 2, 'alto': 3, 'muy_alto': 4},
}

# Libro de códigos para mostrar en la página
LIBRO_CODIGOS = [
    ('Sexo',                  'sexo',                 {1: 'Masculino', 2: 'Femenino'}),
    ('Primera cirugía',       'primera_cirugia',      {1: 'Sí', 2: 'No'}),
    ('Solo biopsia',          'solo_biopsia',         {1: 'Sí', 2: 'No'}),
    ('Extranjero',            'extranjero',           {1: 'Sí', 2: 'No'}),
    ('Provincia nacimiento',  'provincia_nacimiento', {
        1:'Azuay', 2:'Bolívar', 3:'Cañar', 4:'Carchi', 5:'Chimborazo',
        6:'Cotopaxi', 7:'El Oro', 8:'Esmeraldas', 9:'Galápagos', 10:'Guayas',
        11:'Imbabura', 12:'Loja', 13:'Los Ríos', 14:'Manabí', 15:'Morona Santiago',
        16:'Napo', 17:'Orellana', 18:'Pastaza', 19:'Pichincha', 20:'Santa Elena',
        21:'Santo Domingo de los Tsáchilas', 22:'Sucumbíos', 23:'Tungurahua', 24:'Zamora Chinchipe',
    }),
    ('Nivel instrucción',     'nivel_instruccion',    {
        1:'Ninguna', 2:'Primaria Completa', 3:'Primaria Incompleta',
        4:'Secundaria Completa', 5:'Secundaria Incompleta',
        6:'Superior Completa', 7:'Superior Incompleta',
    }),
    ('Etnia',                 'etnia',                {
        1:'Mestizo', 2:'Indígena', 3:'Afroecuatoriano', 4:'Blanco',
        5:'Montubio', 6:'Otro', 7:'No dato',
    }),
    ('Antec. familiares',     'antecedentes_fam',     {1: 'Sí', 2: 'No'}),
    ('Estado',                'estado',               {0: 'Pendiente', 1: 'Completado'}),
    ('GH / Interpretación',   'gh_interp',            {1:'Normal', 2:'Elevado', 3:'Bajo', 4:'No determinado'}),
    ('IGF-1 / Interpretación','igf1_interp',          {1:'Normal', 2:'Elevado', 3:'Bajo', 4:'No determinado'}),
    ('Acromegalia',           'acromegalia',          {1: 'Sí', 2: 'No'}),
    ('Tto. Octreótide',       'tto_previo_octretide', {1: 'Sí', 2: 'No'}),
    ('LH / Interpretación',   'lh_interp',            {1:'Normal', 2:'Elevado', 3:'Bajo', 4:'No determinado'}),
    ('FSH / Interpretación',  'fsh_interp',           {1:'Normal', 2:'Elevado', 3:'Bajo', 4:'No determinado'}),
    ('Estradiol / Interp.',   'estradiol_interp',     {1:'Normal', 2:'Elevado', 3:'Bajo', 4:'No determinado'}),
    ('Testosterona / Interp.','testosterona_interp',  {1:'Normal', 2:'Elevado', 3:'Bajo', 4:'No determinado'}),
    ('Hipogonadismo Fem.',    'hipogonadismo_fem',    {1:'Sí', 2:'No', 3:'No aplica'}),
    ('Hipogonadismo Masc.',   'hipogonadismo_masc',   {1:'Sí', 2:'No', 3:'No aplica'}),
    ('Amenorrea',             'amenorrea',            {1: 'Sí', 2: 'No'}),
    ('Gonadotropinoma',       'gonadotropinoma',      {1: 'Sí', 2: 'No'}),
    ('Hipotiroidismo Central','hipotiroidismo_central',{1: 'Sí', 2: 'No'}),
    ('Tirotropinoma',         'tirotropinoma',        {1: 'Sí', 2: 'No'}),
    ('Cortisol / Interp.',    'cortisol_interp',      {1:'Normal', 2:'Elevado', 3:'Bajo', 4:'No determinado'}),
    ('Eje cort. suficiente',  'eje_corticotropo_suf', {1: 'Sí', 2: 'No'}),
    ('Eje cort. insuficiente','eje_corticotropo_ins', {1: 'Sí', 2: 'No'}),
    ('ACTH / Interpretación', 'acth_interp',          {1:'Normal', 2:'Elevado', 3:'Bajo', 4:'No determinado'}),
    ('Cushing',               'cushing',              {1: 'Sí', 2: 'No'}),
    ('Prolactina / Interp.',  'prolactina_interp',    {1:'Normal', 2:'Baja', 3:'Elevada', 4:'No dato'}),
    ('Tto. Cabergolina',      'tto_previo_cabergolina',{1: 'Sí', 2: 'No'}),
    ('Déficit Vasopresina',   'deficit_vasopresina',  {1: 'Sí', 2: 'No'}),
    ('Apoplejía',             'apoplejia',            {1: 'Sí', 2: 'No'}),
    ('Cefalea',               'cefalea',              {1: 'Sí', 2: 'No'}),
    ('Alt. Campo Visual',     'alt_campo_visual',     {1: 'Sí', 2: 'No'}),
    ('Convulsiones',          'convulsiones',         {1: 'Sí', 2: 'No'}),
    ('Fístula LCR',           'fistula_lcr',          {1: 'Sí', 2: 'No'}),
    ('Sínd. seno cavernoso',  'sind_seno_cavernoso',  {1: 'Sí', 2: 'No'}),
    ('Clasif. tumor',         'tamano_tumor_interp',  {1:'Microadenoma', 2:'Macroadenoma', 3:'Gigante'}),
    ('Invasión seno cav.',    'invasion_seno_cav',    {1: 'Sí', 2: 'No'}),
    ('KNOSP',                 'knosp',                {0:'0', 1:'1', 2:'2', 3:'3', 4:'4'}),
    ('Tipo histológico',      'tipo_histologico',     {
        1:'Prolactinoma', 2:'Somatotropinoma', 3:'Corticotropinoma',
        4:'Gonadotropinoma', 5:'Tirotropinoma', 6:'Carcinoma hipofisario',
        7:'Adenoma no funcionante',
    }),
    ('Reticulina distors.',   'reticulina_distorsionada', {1: 'Sí', 2: 'No'}),
    ('Ki-67 / Interpretación','ki67_interp',          {1:'Muy bajo (<1%)', 2:'Bajo (1-3%)', 3:'Alto (>3%)', 4:'Muy alto (≥10%)'}),
]

# Columnas para el export codificado (mismo orden que exportar_csv pero categóricas → número)
_COLUMNAS_COD = [
    ('nombre',                  lambda p: p.nombre),
    ('hcu',                     lambda p: p.hcu),
    ('fecha_nacimiento',        lambda p: p.fecha_nacimiento.strftime('%d/%m/%Y') if p.fecha_nacimiento else ''),
    ('edad_diagnostico',        lambda p: float(p.edad_diagnostico) if p.edad_diagnostico is not None else ''),
    ('sexo',                    lambda p: _CODIGOS['sexo'].get(p.sexo, '')),
    ('fecha_procedimiento',     lambda p: p.fecha_procedimiento.strftime('%d/%m/%Y') if p.fecha_procedimiento else ''),
    ('procedimiento',           lambda p: p.procedimiento),
    ('cirujano',                lambda p: p.cirujano),
    ('primera_cirugia',         lambda p: _CODIGOS['primera_cirugia'].get(p.primera_cirugia, '')),
    ('solo_biopsia',            lambda p: _CODIGOS['solo_biopsia'].get(p.solo_biopsia, '')),
    ('provincia_nacimiento',    lambda p: _CODIGOS['provincia_nacimiento'].get(p.provincia_nacimiento, '')),
    ('pais_nacimiento',         lambda p: p.pais_nacimiento),
    ('extranjero',              lambda p: _CODIGOS['extranjero'].get(p.extranjero, '')),
    ('nivel_instruccion',       lambda p: _CODIGOS['nivel_instruccion'].get(p.nivel_instruccion, '')),
    ('etnia',                   lambda p: _CODIGOS['etnia'].get(p.etnia, '')),
    ('antecedentes_fam',        lambda p: _CODIGOS['antecedentes_fam'].get(p.antecedentes_fam, '')),
    ('estado',                  lambda p: _CODIGOS['estado'].get(p.estado, '')),
    ('gh',                      lambda p: float(p.gh) if p.gh is not None else ''),
    ('gh_ref_min',              lambda p: float(p.gh_ref_min) if p.gh_ref_min is not None else ''),
    ('gh_ref_max',              lambda p: float(p.gh_ref_max) if p.gh_ref_max is not None else ''),
    ('gh_interp',               lambda p: _CODIGOS['gh_interp'].get(p.gh_interp, '')),
    ('igf1',                    lambda p: float(p.igf1) if p.igf1 is not None else ''),
    ('igf1_ref_min',            lambda p: float(p.igf1_ref_min) if p.igf1_ref_min is not None else ''),
    ('igf1_ref_max',            lambda p: float(p.igf1_ref_max) if p.igf1_ref_max is not None else ''),
    ('igf1_interp',             lambda p: _CODIGOS['igf1_interp'].get(p.igf1_interp, '')),
    ('igf1_index_elevado',      lambda p: float(p.igf1_index_elevado) if p.igf1_index_elevado is not None else ''),
    ('acromegalia',             lambda p: _CODIGOS['acromegalia'].get(p.acromegalia, '')),
    ('tto_previo_octretide',    lambda p: _CODIGOS['tto_previo_octretide'].get(p.tto_previo_octretide, '')),
    ('lh',                      lambda p: float(p.lh) if p.lh is not None else ''),
    ('lh_ref_min',              lambda p: float(p.lh_ref_min) if p.lh_ref_min is not None else ''),
    ('lh_ref_max',              lambda p: float(p.lh_ref_max) if p.lh_ref_max is not None else ''),
    ('lh_interp',               lambda p: _CODIGOS['lh_interp'].get(p.lh_interp, '')),
    ('fsh',                     lambda p: float(p.fsh) if p.fsh is not None else ''),
    ('fsh_ref_min',             lambda p: float(p.fsh_ref_min) if p.fsh_ref_min is not None else ''),
    ('fsh_ref_max',             lambda p: float(p.fsh_ref_max) if p.fsh_ref_max is not None else ''),
    ('fsh_interp',              lambda p: _CODIGOS['fsh_interp'].get(p.fsh_interp, '')),
    ('estradiol',               lambda p: float(p.estradiol) if p.estradiol is not None else ''),
    ('estradiol_ref_min',       lambda p: float(p.estradiol_ref_min) if p.estradiol_ref_min is not None else ''),
    ('estradiol_ref_max',       lambda p: float(p.estradiol_ref_max) if p.estradiol_ref_max is not None else ''),
    ('estradiol_interp',        lambda p: _CODIGOS['estradiol_interp'].get(p.estradiol_interp, '')),
    ('testosterona',            lambda p: float(p.testosterona) if p.testosterona is not None else ''),
    ('testosterona_ref_min',    lambda p: float(p.testosterona_ref_min) if p.testosterona_ref_min is not None else ''),
    ('testosterona_ref_max',    lambda p: float(p.testosterona_ref_max) if p.testosterona_ref_max is not None else ''),
    ('testosterona_interp',     lambda p: _CODIGOS['testosterona_interp'].get(p.testosterona_interp, '')),
    ('hipogonadismo_fem',       lambda p: _CODIGOS['hipogonadismo_fem'].get(p.hipogonadismo_fem, '')),
    ('hipogonadismo_masc',      lambda p: _CODIGOS['hipogonadismo_masc'].get(p.hipogonadismo_masc, '')),
    ('amenorrea',               lambda p: _CODIGOS['amenorrea'].get(p.amenorrea, '')),
    ('gonadotropinoma',         lambda p: _CODIGOS['gonadotropinoma'].get(p.gonadotropinoma, '')),
    ('tsh',                     lambda p: float(p.tsh) if p.tsh is not None else ''),
    ('tsh_ref_min',             lambda p: float(p.tsh_ref_min) if p.tsh_ref_min is not None else ''),
    ('tsh_ref_max',             lambda p: float(p.tsh_ref_max) if p.tsh_ref_max is not None else ''),
    ('t4l',                     lambda p: float(p.t4l) if p.t4l is not None else ''),
    ('t4l_ref_min',             lambda p: float(p.t4l_ref_min) if p.t4l_ref_min is not None else ''),
    ('t4l_ref_max',             lambda p: float(p.t4l_ref_max) if p.t4l_ref_max is not None else ''),
    ('hipotiroidismo_central',  lambda p: _CODIGOS['hipotiroidismo_central'].get(p.hipotiroidismo_central, '')),
    ('tirotropinoma',           lambda p: _CODIGOS['tirotropinoma'].get(p.tirotropinoma, '')),
    ('cortisol',                lambda p: float(p.cortisol) if p.cortisol is not None else ''),
    ('cortisol_ref_min',        lambda p: float(p.cortisol_ref_min) if p.cortisol_ref_min is not None else ''),
    ('cortisol_ref_max',        lambda p: float(p.cortisol_ref_max) if p.cortisol_ref_max is not None else ''),
    ('cortisol_interp',         lambda p: _CODIGOS['cortisol_interp'].get(p.cortisol_interp, '')),
    ('eje_corticotropo_suf',    lambda p: _CODIGOS['eje_corticotropo_suf'].get(p.eje_corticotropo_suf, '')),
    ('eje_corticotropo_ins',    lambda p: _CODIGOS['eje_corticotropo_ins'].get(p.eje_corticotropo_ins, '')),
    ('acth',                    lambda p: float(p.acth) if p.acth is not None else ''),
    ('acth_ref_min',            lambda p: float(p.acth_ref_min) if p.acth_ref_min is not None else ''),
    ('acth_ref_max',            lambda p: float(p.acth_ref_max) if p.acth_ref_max is not None else ''),
    ('acth_interp',             lambda p: _CODIGOS['acth_interp'].get(p.acth_interp, '')),
    ('cushing',                 lambda p: _CODIGOS['cushing'].get(p.cushing, '')),
    ('prolactina',              lambda p: float(p.prolactina) if p.prolactina is not None else ''),
    ('prolactina_ref_min',      lambda p: float(p.prolactina_ref_min) if p.prolactina_ref_min is not None else ''),
    ('prolactina_ref_max',      lambda p: float(p.prolactina_ref_max) if p.prolactina_ref_max is not None else ''),
    ('prolactina_interp',       lambda p: _CODIGOS['prolactina_interp'].get(p.prolactina_interp, '')),
    ('tto_previo_cabergolina',  lambda p: _CODIGOS['tto_previo_cabergolina'].get(p.tto_previo_cabergolina, '')),
    ('deficit_vasopresina',     lambda p: _CODIGOS['deficit_vasopresina'].get(p.deficit_vasopresina, '')),
    ('apoplejia',               lambda p: _CODIGOS['apoplejia'].get(p.apoplejia, '')),
    ('cefalea',                 lambda p: _CODIGOS['cefalea'].get(p.cefalea, '')),
    ('alt_campo_visual',        lambda p: _CODIGOS['alt_campo_visual'].get(p.alt_campo_visual, '')),
    ('convulsiones',            lambda p: _CODIGOS['convulsiones'].get(p.convulsiones, '')),
    ('fistula_lcr',             lambda p: _CODIGOS['fistula_lcr'].get(p.fistula_lcr, '')),
    ('sind_seno_cavernoso',     lambda p: _CODIGOS['sind_seno_cavernoso'].get(p.sind_seno_cavernoso, '')),
    ('tamano_tumor',            lambda p: float(p.tamano_tumor) if p.tamano_tumor is not None else ''),
    ('tamano_tumor_interp',     lambda p: _CODIGOS['tamano_tumor_interp'].get(p.tamano_tumor_interp, '')),
    ('invasion_seno_cav',       lambda p: _CODIGOS['invasion_seno_cav'].get(p.invasion_seno_cav, '')),
    ('knosp',                   lambda p: _CODIGOS['knosp'].get(p.knosp, '')),
    ('tipo_histologico',        lambda p: _CODIGOS['tipo_histologico'].get(p.tipo_histologico, '')),
    ('reticulina_distorsionada',lambda p: _CODIGOS['reticulina_distorsionada'].get(p.reticulina_distorsionada, '')),
    ('ki67',                    lambda p: float(p.ki67) if p.ki67 is not None else ''),
    ('ki67_interp',             lambda p: _CODIGOS['ki67_interp'].get(p.ki67_interp, '')),
    ('observaciones',           lambda p: p.observaciones),
]


@role_required('administrador', 'ver_codificado')
def datos_codificados(request):
    total = Paciente.objects.count()
    return render(request, 'pacientes/datos_codificados.html', {
        'libro_codigos': LIBRO_CODIGOS,
        'total': total,
    })


@role_required('administrador', 'ver_codificado')
def exportar_codificado_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    pacientes = Paciente.objects.all()
    wb = openpyxl.Workbook()

    # ── Hoja 1: Datos codificados ─────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Datos Codificados'

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='1a56db')
    alt_fill = PatternFill('solid', fgColor='F0F4FF')

    for col_idx, (label, _) in enumerate(_COLUMNAS_COD, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True, size=8, color='FFFFFF')
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 30

    for row_idx, p in enumerate(pacientes, start=1):
        for col_idx, (_, fn) in enumerate(_COLUMNAS_COD, start=1):
            try:
                val = fn(p)
            except Exception:
                val = ''
            cell = ws.cell(row=row_idx + 1, column=col_idx, value=val if val != '' else None)
            cell.font = Font(size=8)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col_idx in range(1, len(_COLUMNAS_COD) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # ── Hoja 2: Libro de códigos ──────────────────────────────────────────────
    ws2 = wb.create_sheet('Libro de Códigos')
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 28
    ws2.column_dimensions['C'].width = 8
    ws2.column_dimensions['D'].width = 30

    for c, titulo in enumerate(['Variable', 'Campo', 'Código', 'Etiqueta'], start=1):
        cell = ws2.cell(row=1, column=c, value=titulo)
        cell.font = Font(bold=True, size=9, color='FFFFFF')
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    fila = 2
    for label, campo, codigos in LIBRO_CODIGOS:
        for codigo, etiqueta in sorted(codigos.items()):
            ws2.cell(row=fila, column=1, value=label).font = Font(size=8)
            ws2.cell(row=fila, column=2, value=campo).font = Font(size=8)
            ws2.cell(row=fila, column=3, value=codigo).font = Font(size=8)
            ws2.cell(row=fila, column=4, value=etiqueta).font = Font(size=8)
            for c in range(1, 5):
                ws2.cell(row=fila, column=c).border = border
            fila += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = datetime.now().strftime('%Y%m%d_%H%M')
    resp = HttpResponse(buf.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="datos_codificados_{fecha}.xlsx"'
    return resp


@role_required('administrador', 'ver_codificado')
def exportar_codificado_csv(request):
    import csv
    pacientes = Paciente.objects.all()
    fecha = datetime.now().strftime('%Y%m%d_%H%M')
    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="datos_codificados_{fecha}.csv"'
    resp.write('﻿')
    writer = csv.writer(resp)
    writer.writerow([col for col, _ in _COLUMNAS_COD])
    for p in pacientes:
        fila = []
        for _, fn in _COLUMNAS_COD:
            try:
                fila.append(fn(p))
            except Exception:
                fila.append('')
        writer.writerow(fila)
    return resp
    return render(request, 'pacientes/usuario_confirmar_eliminar.html', {'usuario': usuario})
