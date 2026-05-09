"""
referencias/excel_importer.py
Importador de valores de referencia desde Excel (.xlsx / .xls).

Formato esperado del Excel (una hoja por hormona O todas en la misma hoja):

  Columna A — hormona     : nombre o código (gh, igf1, tsh, ft4, …) OR el título de sección
  Columna B — version     : versión del inserto  (ej. "V1.0, V2.0")
  Columna C — grupo       : grupo clínico / condición
  Columna D — tipo_muestra: tipo de muestra
  Columna E — ref_min     : valor mínimo (número)
  Columna F — ref_max     : valor máximo (número)
  Columna G — unidad      : unidad 1  (ej. "ng/mL")
  Columna H — ref_min2    : mínimo unidad 2  (opcional)
  Columna I — ref_max2    : máximo unidad 2  (opcional)
  Columna J — unidad2     : unidad 2          (opcional)
  Columna K — intervalo_min (opcional)
  Columna L — intervalo_max (opcional)
  Columna M — observaciones (opcional)

La primera fila puede ser cabecera (se detecta automáticamente).
Las filas con hormona vacía se saltan.
"""

import re
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from .models import ValorReferencia, HORMONA_CHOICES

# ── Mapeo de nombres a clave de modelo ────────────────────────────────────────
_ALIAS: dict[str, str] = {
    'gh': 'gh', 'hormona de crecimiento': 'gh', 'hgh': 'gh',
    'igf1': 'igf1', 'igf-1': 'igf1', 'somatomedina': 'igf1', 'igf 1': 'igf1',
    'lh': 'lh', 'luteinizante': 'lh', 'hormona luteinizante': 'lh',
    'fsh': 'fsh', 'foliculo estimulante': 'fsh', 'folículo estimulante': 'fsh',
    'estradiol': 'estradiol', 'e2': 'estradiol',
    'testosterona': 'testosterona', 'testosterone': 'testosterona',
    'tsh': 'tsh', 'tirotropina': 'tsh',
    'ft4': 'ft4', 't4l': 'ft4', 't4 libre': 'ft4', 'tiroxina libre': 'ft4',
    'ft4 - tiroxina libre': 'ft4', 'tiroxina': 'ft4',
    'cortisol': 'cortisol',
    'acth': 'acth', 'adrenocorticotropa': 'acth', 'corticotropina': 'acth',
    'prolactina': 'prolactina', 'prl': 'prolactina',
}

_CLAVES_VALIDAS = {c[0] for c in HORMONA_CHOICES}


def _resolver_hormona(raw: str) -> str | None:
    """Devuelve la clave del modelo o None si no reconoce."""
    if not raw:
        return None
    s = str(raw).strip().lower()
    if s in _CLAVES_VALIDAS:
        return s
    if s in _ALIAS:
        return _ALIAS[s]
    # Búsqueda parcial
    for alias, clave in _ALIAS.items():
        if alias in s or s in alias:
            return clave
    return None


def _decimal(raw) -> Decimal | None:
    if raw is None or str(raw).strip() in ('', '—', '-', 'N/A'):
        return None
    # Extraer el primer número del texto (maneja "11.9 pmol/L")
    m = re.search(r'[-+]?\d+[.,]?\d*', str(raw).replace(',', '.'))
    if not m:
        return None
    try:
        return Decimal(m.group(0))
    except InvalidOperation:
        return None


def _texto(raw) -> str:
    return str(raw).strip() if raw is not None else ''


# ── Cabeceras reconocidas ─────────────────────────────────────────────────────
_HDR_HORMONA   = {'hormona', 'prueba', 'hormone', 'analito'}
_HDR_VERSION   = {'version', 'versión', 'versiones', 'versiones de inserto', 'inserto',
                  'rsiones de inse', 'versiones inserto'}
_HDR_GRUPO     = {'grupo', 'grupo clinico', 'grupo clínico', 'condicion', 'condición',
                  'grupo / condición', 'grupo clínico / condición', 'grupo clinico / condicion'}
_HDR_MUESTRA   = {'tipo muestra', 'tipo de muestra', 'muestra', 'tipo_muestra'}
_HDR_MIN       = {'ref_min', 'ref. min', 'ref. mín', 'valor minimo', 'valor mínimo',
                  'minimo', 'mínimo', 'min', 'mín', 'valor min'}
_HDR_MAX       = {'ref_max', 'ref. max', 'ref. máx', 'valor maximo', 'valor máximo',
                  'maximo', 'máximo', 'max', 'máx', 'valor max'}
_HDR_UNIDAD    = {'unidad', 'unidad 1', 'unidades', 'unit'}
_HDR_MIN2      = {'ref_min2', 'min2', 'mín2', 'valor min 2', 'valor mínimo 2'}
_HDR_MAX2      = {'ref_max2', 'max2', 'máx2', 'valor max 2', 'valor máximo 2'}
_HDR_UNIDAD2   = {'unidad2', 'unidad 2', 'unit2'}
_HDR_INT_MIN   = {'intervalo_min', 'intervalo min', 'int min'}
_HDR_INT_MAX   = {'intervalo_max', 'intervalo max', 'int max'}
_HDR_OBS       = {'observaciones', 'observacion', 'obs', 'notas', 'nota'}


def _detectar_columnas(fila: list) -> dict | None:
    """
    Recibe los valores de la primera fila de cabecera.
    Devuelve un dict {nombre_campo: indice} o None si no parece cabecera.
    """
    mapa = {}
    for i, celda in enumerate(fila):
        if celda is None:
            continue
        s = str(celda).strip().lower()
        if s in _HDR_HORMONA:   mapa['hormona']       = i
        elif s in _HDR_VERSION:  mapa['version']       = i
        elif s in _HDR_GRUPO:    mapa['grupo_clinico'] = i
        elif s in _HDR_MUESTRA:  mapa['tipo_muestra']  = i
        elif s in _HDR_MIN:      mapa.setdefault('ref_min', i)
        elif s in _HDR_MAX:      mapa.setdefault('ref_max', i)
        elif s in _HDR_UNIDAD:   mapa.setdefault('unidad', i)
        elif s in _HDR_MIN2:     mapa['ref_min2']      = i
        elif s in _HDR_MAX2:     mapa['ref_max2']      = i
        elif s in _HDR_UNIDAD2:  mapa['unidad2']       = i
        elif s in _HDR_INT_MIN:  mapa['intervalo_min'] = i
        elif s in _HDR_INT_MAX:  mapa['intervalo_max'] = i
        elif s in _HDR_OBS:      mapa['observaciones'] = i
    return mapa if mapa else None


# ── Función principal ─────────────────────────────────────────────────────────

def importar_excel(archivo, modo: str = 'crear') -> dict:
    """
    Procesa el archivo Excel y crea/actualiza ValorReferencia.

    modo:
      'crear'       — solo crea filas nuevas, omite duplicados
      'actualizar'  — crea y actualiza duplicados (update_or_create)
      'reemplazar'  — elimina todos los registros existentes antes de importar

    Devuelve:
      {'creados': int, 'actualizados': int, 'omitidos': int,
       'errores': [(fila, msg), ...]}
    """
    try:
        wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
    except (InvalidFileException, Exception) as e:
        raise ValueError(f'No se pudo leer el archivo Excel: {e}')

    creados = actualizados = omitidos = 0
    errores: list[tuple[int, str]] = []

    if modo == 'reemplazar':
        ValorReferencia.objects.all().delete()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        filas = list(ws.iter_rows(values_only=True))
        if not filas:
            continue

        # Detectar cabecera
        col = None
        data_start = 0
        for idx, fila in enumerate(filas[:5]):
            col = _detectar_columnas(list(fila))
            if col:
                data_start = idx + 1
                break

        # Sin cabecera → usar orden por defecto (A=hormona, B=version, C=grupo, D=muestra, E=min, F=max, G=unidad)
        if not col:
            col = {
                'hormona': 0, 'version': 1, 'grupo_clinico': 2,
                'tipo_muestra': 3, 'ref_min': 4, 'ref_max': 5, 'unidad': 6,
            }
            data_start = 0

        hormona_contexto = None  # se actualiza con filas de sección/título

        for n_fila, fila in enumerate(filas[data_start:], start=data_start + 2):
            fila = list(fila)
            if all(c is None for c in fila):
                continue

            # Celda hormona
            raw_hormona = fila[col['hormona']] if 'hormona' in col else None
            hormona_k = _resolver_hormona(raw_hormona)

            # ¿Fila de título de sección? (ej. "GH – Hormona de Crecimiento")
            if hormona_k is None and raw_hormona:
                ctx = _resolver_hormona(str(raw_hormona))
                if ctx:
                    hormona_contexto = ctx
                continue

            if hormona_k is None:
                hormona_k = hormona_contexto

            if hormona_k is None:
                omitidos += 1
                continue

            version      = _texto(fila[col['version']]      if 'version'      in col else None)
            grupo        = _texto(fila[col['grupo_clinico']] if 'grupo_clinico' in col else None)
            tipo_muestra = _texto(fila[col['tipo_muestra']]  if 'tipo_muestra'  in col else None)
            unidad       = _texto(fila[col['unidad']]        if 'unidad'        in col else None)
            ref_min      = _decimal(fila[col['ref_min']]     if 'ref_min'       in col else None)
            ref_max      = _decimal(fila[col['ref_max']]     if 'ref_max'       in col else None)
            unidad2      = _texto(fila[col['unidad2']]       if 'unidad2'       in col else None)
            ref_min2     = _decimal(fila[col['ref_min2']]    if 'ref_min2'      in col else None)
            ref_max2     = _decimal(fila[col['ref_max2']]    if 'ref_max2'      in col else None)
            int_min      = _decimal(fila[col['intervalo_min']] if 'intervalo_min' in col else None)
            int_max      = _decimal(fila[col['intervalo_max']] if 'intervalo_max' in col else None)
            obs          = _texto(fila[col['observaciones']] if 'observaciones' in col else None)

            if not version:
                errores.append((n_fila, 'Versión vacía — fila omitida'))
                omitidos += 1
                continue

            defaults = dict(
                tipo_muestra=tipo_muestra,
                unidad=unidad,
                ref_min=ref_min,
                ref_max=ref_max,
                unidad2=unidad2,
                ref_min2=ref_min2,
                ref_max2=ref_max2,
                intervalo_min=int_min,
                intervalo_max=int_max,
                observaciones=obs,
            )

            try:
                if modo == 'crear':
                    _, created = ValorReferencia.objects.get_or_create(
                        hormona=hormona_k, version=version, grupo_clinico=grupo,
                        defaults=defaults,
                    )
                    if created:
                        creados += 1
                    else:
                        omitidos += 1
                else:  # actualizar o reemplazar
                    _, created = ValorReferencia.objects.update_or_create(
                        hormona=hormona_k, version=version, grupo_clinico=grupo,
                        defaults=defaults,
                    )
                    if created:
                        creados += 1
                    else:
                        actualizados += 1
            except Exception as e:
                errores.append((n_fila, str(e)[:120]))
                omitidos += 1

    return {
        'creados': creados,
        'actualizados': actualizados,
        'omitidos': omitidos,
        'errores': errores,
    }
