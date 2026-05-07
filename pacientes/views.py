import io
import json
from datetime import datetime
from functools import wraps
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.core.exceptions import PermissionDenied
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.db.models import Q
from .models import Paciente
from .forms import PacienteForm
from .forms_usuario import UsuarioCrearForm, UsuarioEditarForm


def role_required(*grupos):
    def decorator(view_func):
        @wraps(view_func)
        def wrapped(request, *args, **kwargs):
            if not request.user.is_authenticated:
                from django.contrib.auth.views import redirect_to_login
                return redirect_to_login(request.get_full_path())
            if request.user.is_superuser or request.user.groups.filter(name__in=grupos).exists():
                return view_func(request, *args, **kwargs)
            raise PermissionDenied
        return wrapped
    return decorator


@login_required
def lista(request):
    q = request.GET.get('q', '')
    pacientes = Paciente.objects.all()
    if q:
        pacientes = pacientes.filter(
            Q(nombre__icontains=q) |
            Q(hcu__icontains=q) |
            Q(cirujano__icontains=q) |
            Q(procedimiento__icontains=q)
        )
    return render(request, 'pacientes/lista.html', {'pacientes': pacientes, 'q': q})


@role_required('administrador', 'medico')
def crear(request):
    if request.method == 'POST':
        form = PacienteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro guardado correctamente.')
            return redirect('lista')
    else:
        initial = {k: v for k, v in request.GET.items() if v}
        form = PacienteForm(initial=initial)
    return render(request, 'pacientes/form.html', {'form': form, 'titulo': 'Nuevo Registro'})


@role_required('administrador', 'medico')
def editar(request, pk):
    paciente = get_object_or_404(Paciente, pk=pk)
    form = PacienteForm(request.POST or None, instance=paciente)
    if form.is_valid():
        form.save()
        messages.success(request, 'Registro actualizado correctamente.')
        return redirect('lista')
    return render(request, 'pacientes/form.html', {'form': form, 'titulo': f'Editar — {paciente.nombre}', 'object': paciente})


@role_required('administrador')
def eliminar(request, pk):
    paciente = get_object_or_404(Paciente, pk=pk)
    if request.method == 'POST':
        paciente.delete()
        messages.success(request, 'Registro eliminado.')
        return redirect('lista')
    return render(request, 'pacientes/confirmar_eliminar.html', {'paciente': paciente})


@login_required
def exportar_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    q = request.GET.get('q', '')
    pacientes = Paciente.objects.all()
    if q:
        pacientes = pacientes.filter(
            Q(nombre__icontains=q) | Q(hcu__icontains=q) |
            Q(cirujano__icontains=q) | Q(procedimiento__icontains=q)
        )

    wb = openpyxl.Workbook()

    # ── Hoja 1: Datos Generales ──────────────────────────────────────────────
    _exportar_hoja(wb.active, 'Datos Generales', pacientes, [
        ('N°',                  lambda p, i: i),
        ('Nombre',              lambda p, i: p.nombre),
        ('HCU',                 lambda p, i: p.hcu),
        ('Fecha Nacimiento',    lambda p, i: p.fecha_nacimiento),
        ('Edad Dx',             lambda p, i: float(p.edad_diagnostico) if p.edad_diagnostico is not None else ''),
        ('Sexo',                lambda p, i: p.get_sexo_display()),
        ('Fecha Procedimiento', lambda p, i: p.fecha_procedimiento),
        ('Procedimiento',       lambda p, i: p.procedimiento),
        ('Cirujano',            lambda p, i: p.cirujano),
        ('Primera Cirugía',     lambda p, i: p.get_primera_cirugia_display()),
        ('Solo Biopsia',        lambda p, i: p.get_solo_biopsia_display()),
        ('Provincia',           lambda p, i: p.provincia_nacimiento),
        ('País Nacimiento',     lambda p, i: p.pais_nacimiento),
        ('Extranjero',          lambda p, i: p.get_extranjero_display()),
        ('Nivel Instrucción',   lambda p, i: p.get_nivel_instruccion_display()),
        ('Etnia',               lambda p, i: p.get_etnia_display()),
        ('Antec. Familiares',   lambda p, i: p.get_antecedentes_fam_display()),
        ('Estado',              lambda p, i: p.get_estado_display()),
        ('Registrado',          lambda p, i: p.creado.strftime('%d/%m/%Y') if p.creado else ''),
    ])

    # ── Hoja 2: Hormonas ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Hormonas')
    _exportar_hoja(ws2, 'Hormonas', pacientes, [
        ('N°',                  lambda p, i: i),
        ('Nombre',              lambda p, i: p.nombre),
        ('HCU',                 lambda p, i: p.hcu),
        ('GH',                  lambda p, i: float(p.gh) if p.gh is not None else ''),
        ('GH Ref.Mín',          lambda p, i: float(p.gh_ref_min) if p.gh_ref_min is not None else ''),
        ('GH Ref.Máx',          lambda p, i: float(p.gh_ref_max) if p.gh_ref_max is not None else ''),
        ('GH Interp.',          lambda p, i: p.get_gh_interp_display()),
        ('IGF-1',               lambda p, i: float(p.igf1) if p.igf1 is not None else ''),
        ('IGF-1 Ref.Mín',       lambda p, i: float(p.igf1_ref_min) if p.igf1_ref_min is not None else ''),
        ('IGF-1 Ref.Máx',       lambda p, i: float(p.igf1_ref_max) if p.igf1_ref_max is not None else ''),
        ('IGF-1 Index',         lambda p, i: float(p.igf1_index_elevado) if p.igf1_index_elevado is not None else ''),
        ('Acromegalia',         lambda p, i: p.get_acromegalia_display()),
        ('Tto. Octreótide',     lambda p, i: p.get_tto_previo_octretide_display()),
        ('LH',                  lambda p, i: float(p.lh) if p.lh is not None else ''),
        ('FSH',                 lambda p, i: float(p.fsh) if p.fsh is not None else ''),
        ('Estradiol',           lambda p, i: float(p.estradiol) if p.estradiol is not None else ''),
        ('Testosterona',        lambda p, i: float(p.testosterona) if p.testosterona is not None else ''),
        ('Hipogonadismo Fem.',  lambda p, i: p.get_hipogonadismo_fem_display()),
        ('Hipogonadismo Masc.', lambda p, i: p.get_hipogonadismo_masc_display()),
        ('Amenorrea',           lambda p, i: p.get_amenorrea_display()),
        ('Gonadotropinoma',     lambda p, i: p.get_gonadotropinoma_display()),
        ('TSH',                 lambda p, i: float(p.tsh) if p.tsh is not None else ''),
        ('T4L',                 lambda p, i: float(p.t4l) if p.t4l is not None else ''),
        ('Hipotiroidismo Cent.', lambda p, i: p.get_hipotiroidismo_central_display()),
        ('Tirotropinoma',       lambda p, i: p.get_tirotropinoma_display()),
        ('Cortisol',            lambda p, i: float(p.cortisol) if p.cortisol is not None else ''),
        ('ACTH',                lambda p, i: float(p.acth) if p.acth is not None else ''),
        ('Cushing',             lambda p, i: p.get_cushing_display()),
        ('Prolactina',          lambda p, i: float(p.prolactina) if p.prolactina is not None else ''),
        ('Prolactina Interp.',  lambda p, i: p.get_prolactina_interp_display()),
        ('Tto. Cabergolina',    lambda p, i: p.get_tto_previo_cabergolina_display()),
    ])

    # ── Hoja 3: Tumor e Histología ───────────────────────────────────────────
    ws3 = wb.create_sheet('Tumor e Histología')
    _exportar_hoja(ws3, 'Tumor e Histología', pacientes, [
        ('N°',                  lambda p, i: i),
        ('Nombre',              lambda p, i: p.nombre),
        ('HCU',                 lambda p, i: p.hcu),
        ('Déf. Vasopresina',    lambda p, i: p.get_deficit_vasopresina_display()),
        ('Apoplejía',           lambda p, i: p.get_apoplejia_display()),
        ('Cefalea',             lambda p, i: p.get_cefalea_display()),
        ('Alt. Campo Visual',   lambda p, i: p.get_alt_campo_visual_display()),
        ('Convulsiones',        lambda p, i: p.get_convulsiones_display()),
        ('Fístula LCR',         lambda p, i: p.get_fistula_lcr_display()),
        ('Seno Cavernoso',      lambda p, i: p.get_sind_seno_cavernoso_display()),
        ('Tamaño Tumor (mm)',   lambda p, i: float(p.tamano_tumor) if p.tamano_tumor is not None else ''),
        ('Interp. Tumor',       lambda p, i: p.get_tamano_tumor_interp_display()),
        ('Invasión Seno Cav.',  lambda p, i: p.get_invasion_seno_cav_display()),
        ('KNOSP',               lambda p, i: p.get_knosp_display()),
        ('Tipo Histológico',    lambda p, i: p.get_tipo_histologico_display()),
        ('Reticulina Dist.',    lambda p, i: p.get_reticulina_distorsionada_display()),
        ('Hormonas IHQ',        lambda p, i: p.hormonas_ihq),
        ('Ki-67 (%)',           lambda p, i: float(p.ki67) if p.ki67 is not None else ''),
        ('Ki-67 Interp.',       lambda p, i: p.get_ki67_interp_display()),
        ('Observaciones',       lambda p, i: p.observaciones),
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = datetime.now().strftime('%Y%m%d_%H%M')
    nombre_archivo = f'pacientes_hipofisarios_{fecha}.xlsx'
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return resp


def _exportar_hoja(ws, titulo, pacientes, columnas):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
    title_cell = ws.cell(row=1, column=1, value=titulo)
    title_cell.font = Font(bold=True, size=13, color='FFFFFF')
    title_cell.fill = PatternFill('solid', fgColor='1a56db')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    # Cabecera
    hdr_fill = PatternFill('solid', fgColor='d1e0ff')
    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, (label, _) in enumerate(columnas, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = Font(bold=True, size=9)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 28

    # Datos
    alt_fill = PatternFill('solid', fgColor='F7F9FF')
    for row_idx, p in enumerate(pacientes, start=1):
        excel_row = row_idx + 2
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, (_, fn) in enumerate(columnas, start=1):
            try:
                val = fn(p, row_idx)
            except Exception:
                val = ''
            cell = ws.cell(row=excel_row, column=col_idx, value=val if val != '' else None)
            cell.font = Font(size=9)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if fill:
                cell.fill = fill

    # Ancho automático
    for col_idx in range(1, len(columnas) + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    max_len = max(max_len, len(str(cell.value or '')))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)


@role_required('administrador', 'medico')
@require_POST
def cambiar_estado(request, pk):
    paciente = get_object_or_404(Paciente, pk=pk)
    nuevo = request.POST.get('estado')
    if nuevo not in ('pendiente', 'completado'):
        return JsonResponse({'error': 'Estado inválido'}, status=400)
    paciente.estado = nuevo
    paciente.save(update_fields=['estado', 'actualizado'])
    return JsonResponse({'estado': paciente.estado, 'label': paciente.get_estado_display()})


@role_required('administrador', 'medico')
def subir_pdf(request):
    return render(request, 'pacientes/subir_pdf.html')


@role_required('administrador', 'medico')
@require_POST
def procesar_pdf(request):
    from .pdf_extractor import extraer_todo

    archivo = request.FILES.get('pdf')
    if not archivo:
        return JsonResponse({'error': 'No se recibió ningún archivo.'}, status=400)
    if not archivo.name.lower().endswith('.pdf'):
        return JsonResponse({'error': 'El archivo debe ser un PDF.'}, status=400)
    if archivo.size > 50 * 1024 * 1024:
        return JsonResponse({'error': 'El archivo no debe superar 50 MB.'}, status=400)

    try:
        texto, campos, paginas, gemini_status = extraer_todo(archivo)
        return JsonResponse({
            'campos': campos,
            'paginas': paginas,
            'gemini_status': gemini_status,
            'texto_crudo': texto[:4000],
            'total_campos': len(campos),
        })
    except Exception as e:
        return JsonResponse({'error': f'No se pudo procesar el PDF: {str(e)}'}, status=500)


@login_required
def imprimir(request, pk):
    paciente = get_object_or_404(Paciente, pk=pk)
    return render(request, 'pacientes/imprimir.html', {'paciente': paciente})


# ── Gestión de usuarios (solo administrador) ──────────────────────────────────

@role_required('administrador')
def usuarios_lista(request):
    usuarios = User.objects.exclude(is_superuser=True).order_by('username')
    return render(request, 'pacientes/usuarios_lista.html', {'usuarios': usuarios})


@role_required('administrador')
def usuario_crear(request):
    form = UsuarioCrearForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        cd = form.cleaned_data
        u = User.objects.create_user(
            username=cd['username'],
            password=cd['password1'],
            first_name=cd.get('first_name', ''),
            last_name=cd.get('last_name', ''),
            email=cd.get('email', ''),
        )
        if cd['rol']:
            grupo, _ = Group.objects.get_or_create(name=cd['rol'])
            u.groups.set([grupo])
        messages.success(request, f'Usuario "{u.username}" creado correctamente.')
        return redirect('usuarios_lista')
    return render(request, 'pacientes/usuario_form.html', {
        'form': form, 'titulo': 'Nuevo Usuario', 'es_edicion': False,
    })


@role_required('administrador')
def usuario_editar(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    rol_actual = usuario.groups.values_list('name', flat=True).first() or ''
    initial = {
        'username': usuario.username,
        'first_name': usuario.first_name,
        'last_name': usuario.last_name,
        'email': usuario.email,
        'rol': rol_actual,
        'is_active': usuario.is_active,
    }
    form = UsuarioEditarForm(request.POST or None, usuario=usuario, initial=initial)
    if request.method == 'POST' and form.is_valid():
        cd = form.cleaned_data
        usuario.username   = cd['username']
        usuario.first_name = cd.get('first_name', '')
        usuario.last_name  = cd.get('last_name', '')
        usuario.email      = cd.get('email', '')
        usuario.is_active  = cd.get('is_active', True)
        if cd.get('password1'):
            usuario.set_password(cd['password1'])
        usuario.save()
        if cd['rol']:
            grupo, _ = Group.objects.get_or_create(name=cd['rol'])
            usuario.groups.set([grupo])
        else:
            usuario.groups.clear()
        messages.success(request, f'Usuario "{usuario.username}" actualizado.')
        return redirect('usuarios_lista')
    return render(request, 'pacientes/usuario_form.html', {
        'form': form, 'titulo': f'Editar — {usuario.username}',
        'es_edicion': True, 'usuario': usuario,
    })


@role_required('administrador')
def usuario_eliminar(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    if usuario.is_superuser:
        messages.error(request, 'No se puede eliminar a un superusuario.')
        return redirect('usuarios_lista')
    if request.method == 'POST':
        nombre = usuario.username
        usuario.delete()
        messages.success(request, f'Usuario "{nombre}" eliminado.')
        return redirect('usuarios_lista')
    return render(request, 'pacientes/usuario_confirmar_eliminar.html', {'usuario': usuario})
